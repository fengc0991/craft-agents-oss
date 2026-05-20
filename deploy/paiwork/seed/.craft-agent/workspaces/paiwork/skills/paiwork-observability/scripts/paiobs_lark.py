#!/usr/bin/env python3.11
"""Publish PaiWork observability reports to Feishu/Lark."""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess
import sys
import time
import uuid
from pathlib import Path
from typing import Any

import paiobs

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - deployment image normally includes openpyxl.
    load_workbook = None


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
SKILL_LARK_CLI = SKILL_DIR / "bin" / "lark-cli"
SKILL_LARK_HOME = SKILL_DIR / ".lark-cli-home"
SKILL_FEISHU_SYNC = SKILL_DIR / "bin" / "feishu-sync"
DEFAULT_CONTAINER = "fastgpt-sandbox-shell-http"
DEFAULT_LARK_PROFILE = "enterprise-fengchao"
REQUIRED_SEND_SCOPES = ["im:message.send_as_user", "im:message"]
OPTIONAL_CONTACT_SCOPES = ["contact:user:search"]

# Responsibility owner name -> Feishu open_id mapping.
# Loaded from PAI_OBS_OWNER_OPEN_IDS env var (JSON object) or hardcoded below.
# When configured, 责任人 is rendered as a "people" type field with blue @ mentions.
# Format: {"@王能": "ou_xxx", "@凤超": "ou_yyy", ...}
_DEFAULT_OWNER_OPEN_IDS: dict[str, str] = {
    "@凤超": "ou_d6a21aa8610840878e145f22ed96d866",
}


def _load_owner_open_ids() -> dict[str, str]:
    env_value = os.environ.get("PAI_OBS_OWNER_OPEN_IDS", "").strip()
    if env_value:
        try:
            parsed = json.loads(env_value)
            if isinstance(parsed, dict):
                return {str(k): str(v) for k, v in parsed.items() if str(v).startswith("ou_")}
        except (json.JSONDecodeError, TypeError):
            pass
    return dict(_DEFAULT_OWNER_OPEN_IDS)


RESPONSIBILITY_OWNER_OPEN_IDS = _load_owner_open_ids()
FOCUS_TASK_URL_FIELDS = [
    "query链接",
    "最终文件产物",
]
FOCUS_TASK_SINGLE_SELECT_FIELDS = [
    "查询一级分类",
    "查询二级分类",
    "问题一级分类",
    "问题二级分类",
    "用户类型",
    "用户角色",
    "产品类型",
]
FOCUS_TASK_TEXT_FIELDS = [
    "关注类型",
    "关注原因",
    "关注阈值",
    "query内容",
    "answer内容",
    "错误信息",
    "低分原因",
    "改进建议",
    "证据",
    "skills",
    "用户ID",
    "用户名",
    "用户机构",
    "机构性质",
    "机构类型",
    "queryid",
    "session_id",
    "问题id",
    "责任人",
]


class LarkError(RuntimeError):
    pass


def run_command(cmd: list[str], cwd: str | Path | None = None, env: dict[str, str] | None = None) -> tuple[str, str]:
    process = subprocess.Popen(
        cmd,
        cwd=str(cwd) if cwd else None,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )
    stdout, stderr = process.communicate()
    if process.returncode != 0:
        detail = " | ".join(part.strip() for part in [stdout, stderr] if part.strip())
        raise LarkError(f"command failed ({process.returncode}): {' '.join(cmd)} :: {detail}")
    return stdout, stderr


def find_json_payload(text: str) -> Any:
    decoder = json.JSONDecoder()
    for index, char in enumerate(text):
        if char not in "[{":
            continue
        try:
            payload, _end = decoder.raw_decode(text[index:])
            return payload
        except ValueError:
            continue
    raise LarkError(f"command did not return JSON: {text.strip()[:500]}")


def feishu_sync_bin() -> str:
    env_value = os.environ.get("PAI_OBS_FEISHU_SYNC") or os.environ.get("FEISHU_SYNC_BIN") or ""
    if env_value:
        return env_value
    if SKILL_FEISHU_SYNC.is_file():
        return str(SKILL_FEISHU_SYNC)
    found = shutil.which("feishu-sync")
    if found:
        return found
    raise LarkError("legacy feishu-sync not found; direct skill-local lark-cli publishing is expected")


def feishu_sync_json(args: list[str]) -> dict[str, Any]:
    stdout, stderr = run_command([feishu_sync_bin(), *args, "--format", "json"])
    payload = find_json_payload(stdout or stderr)
    if not isinstance(payload, dict):
        raise LarkError("feishu-sync returned non-object JSON")
    return payload


def lark_cli_environment() -> dict[str, str]:
    env = os.environ.copy()
    home = Path(env.get("PAI_OBS_LARK_HOME") or env.get("LARK_CLI_HOME") or SKILL_LARK_HOME).resolve()
    home.mkdir(parents=True, exist_ok=True)
    env["LARK_CLI_HOME"] = str(home)
    env["HOME"] = str(home)
    return env


def bool_env(name: str, default: bool = False) -> bool:
    value = os.environ.get(name, "").strip().lower()
    if not value:
        return default
    return value in {"1", "true", "yes", "on"}


def lark_cli_invocation(cwd: str | Path | None = None) -> tuple[list[str], str | None, dict[str, str] | None, str]:
    env_bin = os.environ.get("PAI_OBS_LARK_CLI", "").strip()
    candidates = [Path(env_bin)] if env_bin else []
    candidates.append(SKILL_LARK_CLI)
    found = shutil.which("lark-cli")
    if found:
        candidates.append(Path(found))
    for candidate in candidates:
        if candidate.is_file() and os.access(candidate, os.X_OK):
            return [str(candidate)], str(cwd) if cwd else None, lark_cli_environment(), "local"
    docker = shutil.which("docker")
    if docker and bool_env("PAI_OBS_LARK_ALLOW_DOCKER"):
        return [
            docker,
            "exec",
            "-w",
            str(cwd or "/workspace"),
            os.environ.get("PAI_OBS_LARK_CONTAINER", DEFAULT_CONTAINER),
            "lark-cli",
        ], None, None, "docker"
    raise LarkError(f"lark-cli not found in skill or PATH; expected executable at {SKILL_LARK_CLI}")


def lark_cli_prefix() -> list[str]:
    return lark_cli_invocation()[0]


def lark_json(args: list[str], cwd: str | Path | None = None) -> dict[str, Any]:
    profile = os.environ.get("PAI_OBS_LARK_PROFILE", "").strip()
    profile_args = ["--profile", profile] if profile else []
    prefix, run_cwd, env, _mode = lark_cli_invocation(cwd)
    stdout, stderr = run_command([*prefix, *profile_args, *args], cwd=run_cwd, env=env)
    payload = find_json_payload(stdout or stderr)
    if not isinstance(payload, dict):
        raise LarkError("lark-cli returned non-object JSON")
    return payload


def direct_lark_enabled() -> bool:
    value = os.environ.get("PAI_OBS_LARK_DIRECT", "").strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    return True


def prepare_file_for_lark(file_path: str | Path) -> tuple[str | Path, str]:
    source = Path(file_path).resolve()
    if not source.is_file():
        raise LarkError(f"file not found: {source}")
    _prefix, _cwd, _env, mode = lark_cli_invocation()
    if mode == "local":
        return source.parent, source.name
    docker = shutil.which("docker")
    if not docker:
        raise LarkError("docker is required to pass local files to containerized lark-cli")
    container = os.environ.get("PAI_OBS_LARK_CONTAINER", DEFAULT_CONTAINER)
    remote_dir = f"/tmp/paiobs_lark_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    remote_name = source.name
    run_command([docker, "exec", container, "mkdir", "-p", remote_dir])
    run_command([docker, "cp", str(source), f"{container}:{remote_dir}/{remote_name}"])
    return remote_dir, remote_name


def walk_json(value: Any):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from walk_json(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk_json(child)


def first_user_id(payload: Any) -> str:
    for node in walk_json(payload):
        if not isinstance(node, dict):
            continue
        for key in ("open_id", "openId", "user_id", "userId"):
            value = str(node.get(key) or "").strip()
            if value.startswith("ou_"):
                return value
    return ""


def auth_status_user_id(query: str) -> str:
    try:
        payload = lark_json(["auth", "status"])
    except LarkError:
        return ""
    if not isinstance(payload, dict):
        return ""
    query_norm = query.strip().lower()
    user_name = str(payload.get("userName") or payload.get("user_name") or "").strip().lower()
    user_id = str(payload.get("userOpenId") or payload.get("user_open_id") or payload.get("open_id") or "").strip()
    if user_id.startswith("ou_") and (not query_norm or query_norm == user_name):
        return user_id
    return ""


def resolve_user_id(query: str) -> str:
    try:
        payload = lark_json(["contact", "+search-user", "--as", "user", "--query", query, "--format", "json"])
        user_id = first_user_id(payload)
    except LarkError:
        user_id = auth_status_user_id(query)
        if user_id:
            return user_id
        raise
    if not user_id:
        user_id = auth_status_user_id(query)
    if not user_id:
        raise LarkError(f"could not resolve Feishu user for query: {query}")
    return user_id


def create_doc(report_md: str, doc_path: str, dry_run: bool = False) -> dict[str, Any]:
    if dry_run:
        return {"dry_run": True, "action": "create-doc", "path": doc_path, "source": report_md}
    if direct_lark_enabled():
        cwd, name = prepare_file_for_lark(report_md)
        title = Path(doc_path).stem or Path(report_md).stem or "paiwork_observability_report"
        payload = lark_json(["docs", "+create", "--as", "user", "--title", title, "--markdown", f"@{name}"], cwd=cwd)
        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        return {
            "action": "create-doc",
            "mode": "direct-lark",
            "path": doc_path,
            "source": str(Path(report_md).resolve()),
            "url": str(data.get("doc_url") or data.get("url") or ""),
            "remote_url": str(data.get("doc_url") or data.get("url") or ""),
            "remote_token": str(data.get("doc_id") or data.get("docId") or ""),
            "payload": payload,
        }
    return feishu_sync_json(["create-doc", doc_path, "--from-file", str(Path(report_md).resolve())])


def first_bitable_table_id(base_token: str) -> str:
    table_ids = bitable_table_ids(base_token)
    return table_ids[0] if table_ids else ""


def bitable_tables(base_token: str) -> list[dict[str, str]]:
    payload = lark_json(["base", "+table-list", "--as", "user", "--base-token", base_token])
    data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
    tables = data.get("tables") if isinstance(data.get("tables"), list) else []
    result = []
    for table in tables:
        if not isinstance(table, dict) or not table.get("id"):
            continue
        result.append(
            {
                "id": str(table.get("id") or ""),
                "name": str(table.get("name") or table.get("title") or ""),
            }
        )
    return result


def bitable_table_ids(base_token: str) -> list[str]:
    return [table["id"] for table in bitable_tables(base_token) if table.get("id")]


def update_bitable_text_fields(base_token: str, table_id: str, field_names: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    return update_bitable_fields(base_token, table_id, {field_name: "text" for field_name in field_names})


def update_bitable_url_fields(base_token: str, table_id: str, field_names: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    return update_bitable_fields(base_token, table_id, {field_name: "url" for field_name in field_names})


def update_bitable_single_select_fields(
    base_token: str,
    table_id: str,
    field_names: list[str],
    options_by_field: dict[str, list[str]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    # Note: lark-cli <=1.0.x does not support 'property' in field-update JSON.
    # We set the field type to single_select first, then the select values are
    # populated via restore_bitable_select_values (record batch update) which
    # auto-creates options when writing values.
    return update_bitable_fields(base_token, table_id, {fn: "single_select" for fn in field_names})


def csv_fieldnames(csv_path: str) -> set[str]:
    try:
        with Path(csv_path).open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            return {str(field or "").strip() for field in next(reader, []) if str(field or "").strip()}
    except OSError:
        return set()


def workbook_fieldnames(path: str) -> set[str]:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return csv_fieldnames(path)
    if suffix not in {".xlsx", ".xlsm"} or load_workbook is None:
        return set()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return set()
    fields: set[str] = set()
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            for row in rows:
                fields.update(str(value or "").strip() for value in row if str(value or "").strip())
                break
    finally:
        workbook.close()
    return fields


def workbook_rows_by_sheet(path: str, field_names: list[str]) -> dict[str, list[dict[str, Any]]]:
    suffix = Path(path).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"} or load_workbook is None:
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    result: dict[str, list[dict[str, Any]]] = {}
    needed = {"queryid", "session_id", "task_index", *field_names}
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(rows)]
            except StopIteration:
                result[worksheet.title] = []
                continue
            sheet_rows = []
            for row in rows:
                item = {
                    header: row[index] if index < len(row) else ""
                    for index, header in enumerate(headers)
                    if header in needed
                }
                if any(item.get(field) not in (None, "", [], {}) for field in needed):
                    sheet_rows.append(item)
            result[worksheet.title] = sheet_rows
    finally:
        workbook.close()
    return result


def workbook_select_options(path: str, field_names: list[str]) -> dict[str, list[str]]:
    """Collect all unique non-empty values per field across all sheets of the workbook."""
    suffix = Path(path).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"} or load_workbook is None:
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    options: dict[str, dict[str, None]] = {field: {} for field in field_names}
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(rows)]
            except StopIteration:
                continue
            field_indices = {}
            for index, header in enumerate(headers):
                if header in options:
                    field_indices[header] = index
            for row in rows:
                for field_name, col_index in field_indices.items():
                    if col_index < len(row):
                        value = str(row[col_index] or "").strip()
                        if value:
                            options[field_name][value] = None
    finally:
        workbook.close()
    return {field: list(vals.keys()) for field, vals in options.items() if vals}


def normalized_record_value(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def record_match_key(row: dict[str, Any]) -> tuple[str, str]:
    query_id = normalized_record_value(row.get("queryid"))
    if query_id:
        return "queryid", query_id
    session_id = normalized_record_value(row.get("session_id"))
    task_index = normalized_record_value(row.get("task_index"))
    if session_id and task_index:
        return "session_task", f"{session_id}:{task_index}"
    return "", ""


def jsonable_record_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def chunked(items: list[str], size: int) -> list[list[str]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def bitable_record_rows(base_token: str, table_id: str) -> list[dict[str, Any]]:
    records = []
    offset = 0
    while True:
        payload = lark_json(
            [
                "base",
                "+record-list",
                "--as",
                "user",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--format",
                "json",
                "--limit",
                "500",
                "--offset",
                str(offset),
            ]
        )
        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        field_names = data.get("fields") if isinstance(data.get("fields"), list) else []
        rows = data.get("data") if isinstance(data.get("data"), list) else []
        record_ids = data.get("record_id_list") if isinstance(data.get("record_id_list"), list) else []
        for record_id, row in zip(record_ids, rows):
            if not isinstance(row, list):
                continue
            fields = {
                str(field_names[index]): row[index] if index < len(row) else None
                for index in range(len(field_names))
            }
            records.append({"record_id": str(record_id or ""), "fields": fields})
        if not data.get("has_more"):
            break
        offset += len(rows)
        if not rows:
            break
    return records


def restore_bitable_select_values(
    base_token: str,
    table_id: str,
    sheet_rows: list[dict[str, Any]],
    field_names: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    desired_by_key = {
        key: row
        for row in sheet_rows
        for key in [record_match_key(row)]
        if key[0]
    }
    updates: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    if not desired_by_key:
        return updates, errors
    grouped: dict[str, dict[str, Any]] = {}
    for record in bitable_record_rows(base_token, table_id):
        record_id = str(record.get("record_id") or "")
        fields = record.get("fields") if isinstance(record.get("fields"), dict) else {}
        if not record_id:
            continue
        source = desired_by_key.get(record_match_key(fields))
        if source is None:
            continue
        patch = {
            field_name: jsonable_record_value(source.get(field_name))
            for field_name in field_names
            if source.get(field_name) not in (None, "", [], {})
        }
        if not patch:
            continue
        group_key = json.dumps(patch, ensure_ascii=False, sort_keys=True)
        group = grouped.setdefault(group_key, {"patch": patch, "record_ids": []})
        group["record_ids"].append(record_id)
    for group in grouped.values():
        patch = group["patch"]
        record_ids = group["record_ids"]
        for batch_ids in chunked(record_ids, 500):
            batch_json = {"record_id_list": batch_ids, "patch": patch}
            try:
                payload = lark_json(
                    [
                        "base",
                        "+record-batch-update",
                        "--as",
                        "user",
                        "--base-token",
                        base_token,
                        "--table-id",
                        table_id,
                        "--json",
                        json.dumps(batch_json, ensure_ascii=False),
                    ]
                )
                updates.append({"record_ids": batch_ids, "count": len(batch_ids), "fields": sorted(patch), "payload": payload})
            except LarkError as exc:
                errors.append({"record_ids": batch_ids, "count": len(batch_ids), "fields": sorted(patch), "error": str(exc)})
    return updates, errors


def update_bitable_fields(base_token: str, table_id: str, field_types: dict[str, str]) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    specs = {field_name: {"name": field_name, "type": field_type} for field_name, field_type in field_types.items()}
    return update_bitable_fields_with_spec(base_token, table_id, specs)


def update_bitable_fields_with_spec(
    base_token: str,
    table_id: str,
    field_specs: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    updates: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    for index, (field_name, spec) in enumerate(field_specs.items()):
        if index > 0:
            time.sleep(0.3)  # rate limit mitigation for Feishu field-update API
        field_type = spec.get("type", "text")
        try:
            payload = lark_json(
                [
                    "base",
                    "+field-update",
                    "--as",
                    "user",
                    "--yes",
                    "--base-token",
                    base_token,
                    "--table-id",
                    table_id,
                    "--field-id",
                    field_name,
                    "--json",
                    json.dumps(spec, ensure_ascii=False),
                ]
            )
            updates.append({"field": field_name, "type": field_type, "payload": payload})
        except LarkError as exc:
            text = str(exc)
            if "no operation produced" in text:
                updates.append(
                    {
                        "field": field_name,
                        "type": field_type,
                        "payload": {"ok": True, "no_op": True, "message": f"field already matched desired {field_type} type"},
                    }
                )
            else:
                errors.append({"field": field_name, "error": text})
    return updates, errors


def _parse_owner_names(text: str) -> list[str]:
    """Parse owner text like '@王能' or '@王宝涵 @王意荃' into individual names with @."""
    import re
    return re.findall(r'@\S+', text.strip())


def _resolve_owner_open_ids(owner_text: str) -> list[str]:
    """Resolve owner text to a list of open_ids using RESPONSIBILITY_OWNER_OPEN_IDS."""
    names = _parse_owner_names(owner_text)
    ids = []
    for name in names:
        open_id = RESPONSIBILITY_OWNER_OPEN_IDS.get(name)
        if open_id:
            ids.append(open_id)
    return ids


def update_bitable_people_field(
    base_token: str,
    table_id: str,
    field_name: str,
    sheet_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    """Convert a text field to people type and batch-update with structured user references."""
    updates: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    # Step 1: Convert field to user type with multiple=true
    # Note: lark-cli <=1.0.x may not support 'property' in field-update.
    # Try with property first, fall back to plain type.
    spec = {"name": field_name, "type": "user"}
    field_updates, field_errors = update_bitable_fields_with_spec(base_token, table_id, {field_name: spec})
    updates.extend(field_updates)
    errors.extend(field_errors)
    if field_errors:
        return updates, errors
    # Step 2: Build record patches - map each record's owner text to open_id list
    desired_by_key: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in sheet_rows:
        owner_text = str(row.get(field_name) or "").strip()
        if not owner_text:
            continue
        open_ids = _resolve_owner_open_ids(owner_text)
        if not open_ids:
            continue
        key = record_match_key(row)
        if key[0]:
            desired_by_key[key] = [{"id": oid} for oid in open_ids]
    if not desired_by_key:
        return updates, errors
    # Step 3: Match records and batch update
    grouped: dict[str, list[str]] = {}
    for record in bitable_record_rows(base_token, table_id):
        record_id = str(record.get("record_id") or "")
        fields = record.get("fields") if isinstance(record.get("fields"), dict) else {}
        if not record_id:
            continue
        people_value = desired_by_key.get(record_match_key(fields))
        if people_value is None:
            continue
        group_key = json.dumps(people_value, ensure_ascii=False, sort_keys=True)
        grouped.setdefault(group_key, []).append(record_id)
    for group_key, record_ids in grouped.items():
        people_value = json.loads(group_key)
        for batch_ids in chunked(record_ids, 500):
            batch_json = {"record_id_list": batch_ids, "patch": {field_name: people_value}}
            try:
                payload = lark_json(
                    [
                        "base",
                        "+record-batch-update",
                        "--as",
                        "user",
                        "--base-token",
                        base_token,
                        "--table-id",
                        table_id,
                        "--json",
                        json.dumps(batch_json, ensure_ascii=False),
                    ]
                )
                updates.append({"field": field_name, "record_ids": batch_ids, "count": len(batch_ids), "payload": payload})
            except LarkError as exc:
                errors.append({"field": field_name, "record_ids": batch_ids, "count": len(batch_ids), "error": str(exc)})
    return updates, errors


def create_bitable(table_path: str, bitable_path: str, dry_run: bool = False) -> dict[str, Any]:
    if dry_run:
        return {"dry_run": True, "action": "create-bitable", "path": bitable_path, "source": table_path}
    if direct_lark_enabled():
        cwd, name = prepare_file_for_lark(table_path)
        title = Path(bitable_path).stem or Path(table_path).stem or "paiwork_focus_tasks"
        payload = lark_json(["drive", "+import", "--as", "user", "--file", name, "--name", title, "--type", "bitable"], cwd=cwd)
        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        base_token = str(data.get("token") or "")
        tables = bitable_tables(base_token) if base_token else []
        table_ids = [table["id"] for table in tables if table.get("id")]
        field_updates: list[dict[str, Any]] = []
        field_update_errors: list[dict[str, str]] = []
        select_value_updates: list[dict[str, Any]] = []
        select_value_update_errors: list[dict[str, str]] = []
        people_updates: list[dict[str, Any]] = []
        people_update_errors: list[dict[str, str]] = []
        if base_token and table_ids:
            headers = workbook_fieldnames(table_path)
            select_rows_by_sheet = workbook_rows_by_sheet(table_path, FOCUS_TASK_SINGLE_SELECT_FIELDS)
            select_options = workbook_select_options(table_path, FOCUS_TASK_SINGLE_SELECT_FIELDS)
            # Also read rows for people field (责任人)
            people_field_name = "责任人"
            people_rows_by_sheet = workbook_rows_by_sheet(table_path, [people_field_name]) if RESPONSIBILITY_OWNER_OPEN_IDS else {}
            text_fields = [field for field in FOCUS_TASK_TEXT_FIELDS if not headers or field in headers]
            # If we have owner open_ids, remove 责任人 from text fields (it'll be people type)
            if RESPONSIBILITY_OWNER_OPEN_IDS and people_field_name in text_fields:
                text_fields = [f for f in text_fields if f != people_field_name]
            url_fields = [field for field in FOCUS_TASK_URL_FIELDS if not headers or field in headers]
            single_select_fields = [field for field in FOCUS_TASK_SINGLE_SELECT_FIELDS if not headers or field in headers]
            for table_index, table in enumerate(tables):
                if table_index > 0:
                    time.sleep(1.0)  # rate limit mitigation between tables
                table_id = table["id"]
                text_updates, text_errors = update_bitable_text_fields(base_token, table_id, text_fields)
                url_updates, url_errors = update_bitable_url_fields(base_token, table_id, url_fields)
                select_updates, select_errors = update_bitable_single_select_fields(
                    base_token, table_id, single_select_fields, options_by_field=select_options,
                )
                field_updates.extend({**item, "table_id": table_id} for item in [*text_updates, *url_updates, *select_updates])
                field_update_errors.extend({**item, "table_id": table_id} for item in [*text_errors, *url_errors, *select_errors])
                if single_select_fields and select_rows_by_sheet:
                    sheet_name = table.get("name") or ""
                    sheet_rows = select_rows_by_sheet.get(sheet_name)
                    if sheet_rows is None and table_index < len(select_rows_by_sheet):
                        sheet_rows = list(select_rows_by_sheet.values())[table_index]
                    value_updates, value_errors = restore_bitable_select_values(
                        base_token,
                        table_id,
                        sheet_rows or [],
                        single_select_fields,
                    )
                    select_value_updates.extend({**item, "table_id": table_id} for item in value_updates)
                    select_value_update_errors.extend({**item, "table_id": table_id} for item in value_errors)
                # Update 责任人 to people type if owner open_ids are configured
                if RESPONSIBILITY_OWNER_OPEN_IDS and people_rows_by_sheet:
                    sheet_name = table.get("name") or ""
                    people_rows = people_rows_by_sheet.get(sheet_name)
                    if people_rows is None and table_index < len(people_rows_by_sheet):
                        people_rows = list(people_rows_by_sheet.values())[table_index]
                    p_updates, p_errors = update_bitable_people_field(
                        base_token, table_id, people_field_name, people_rows or [],
                    )
                    people_updates.extend({**item, "table_id": table_id} for item in p_updates)
                    people_update_errors.extend({**item, "table_id": table_id} for item in p_errors)
        return {
            "action": "create-bitable",
            "mode": "direct-lark",
            "path": bitable_path,
            "source": str(Path(table_path).resolve()),
            "url": str(data.get("url") or ""),
            "remote_url": str(data.get("url") or ""),
            "remote_token": base_token,
            "table_id": table_ids[0] if table_ids else "",
            "table_ids": table_ids,
            "field_updates": field_updates,
            "field_update_errors": field_update_errors,
            "select_value_updates": select_value_updates,
            "select_value_update_errors": select_value_update_errors,
            "people_updates": people_updates,
            "people_update_errors": people_update_errors,
            "payload": payload,
        }
    if Path(table_path).suffix.lower() != ".csv":
        raise LarkError("multi-sheet focus workbook requires direct lark import; set --lark-profile or PAI_OBS_LARK_PROFILE")
    return feishu_sync_json(["create-bitable", bitable_path, "--csv", str(Path(table_path).resolve())])


def send_with_identity(args: list[str], *, send_as: str, dry_run_payload: dict[str, Any] | None = None, cwd: str | Path | None = None) -> dict[str, Any]:
    if dry_run_payload is not None:
        return dry_run_payload
    identities = ["user", "bot"] if send_as == "auto" else [send_as]
    last_error: LarkError | None = None
    for identity in identities:
        try:
            payload = lark_json(["im", "+messages-send", "--as", identity, *args], cwd=cwd)
            payload.setdefault("send_as", identity)
            return payload
        except LarkError as exc:
            last_error = exc
            if identity != identities[-1]:
                continue
            raise
    raise last_error or LarkError("message send failed")


def send_message(user_id: str, markdown: str, dry_run: bool = False, send_as: str = "auto") -> dict[str, Any]:
    if dry_run:
        return {"dry_run": True, "action": "send-message", "user_id": user_id, "markdown": markdown}
    return send_with_identity(["--user-id", user_id, "--markdown", markdown], send_as=send_as)


def send_file(user_id: str, file_path: str, dry_run: bool = False, send_as: str = "auto") -> dict[str, Any]:
    if dry_run:
        return {"dry_run": True, "action": "send-file", "user_id": user_id, "file": file_path}
    resolved = Path(file_path).resolve()
    return send_with_identity(["--user-id", user_id, "--file", resolved.name], send_as=send_as, cwd=str(resolved.parent))


def render_message_template(template: str, values: dict[str, Any]) -> str:
    rendered = template
    for key, value in values.items():
        rendered = rendered.replace("{" + key + "}", str(value or ""))
    return rendered


def append_missing_message_links(message: str, *, doc_url: str = "", bitable_url: str = "") -> str:
    additions = []
    if doc_url and doc_url not in message:
        additions.append(f"报告：{doc_url}")
    if bitable_url and bitable_url not in message:
        additions.append(f"关注任务表：{bitable_url}")
    if not additions:
        return message
    base = message.rstrip()
    return "\n".join([base, *additions] if base else additions)


def split_scopes(value: Any) -> set[str]:
    return {item.strip() for item in str(value or "").replace(",", " ").split() if item.strip()}


def default_doc_path(report_md: str) -> str:
    stem = Path(report_md).stem or "paiwork_observability_report"
    return f"paiwork_reports/{stem}.md"


def default_bitable_path(csv_path: str) -> str:
    stem = Path(csv_path).stem or "paiwork_focus_tasks"
    return f"paiwork_reports/{stem}.bitable.xlsx"


def cmd_publish(args: argparse.Namespace) -> None:
    if args.lark_profile:
        os.environ["PAI_OBS_LARK_PROFILE"] = args.lark_profile
    if not Path(args.report_md).is_file():
        raise SystemExit(f"report markdown not found: {args.report_md}")
    focus_task_file = args.focus_task_file or args.low_score_csv
    if focus_task_file and not Path(focus_task_file).is_file():
        raise SystemExit(f"focus task table file not found: {focus_task_file}")
    doc_path = args.doc_path or default_doc_path(args.report_md)
    bitable_path = args.bitable_path or (default_bitable_path(focus_task_file) if focus_task_file else "")
    publish_errors: list[dict[str, str]] = []
    try:
        doc_result = create_doc(args.report_md, doc_path, dry_run=args.dry_run)
    except LarkError as exc:
        if not args.fallback_files:
            raise
        publish_errors.append({"stage": "create_doc", "error": str(exc)})
        doc_result = {"ok": False, "error": str(exc), "fallback": "file"}
    try:
        bitable_result = create_bitable(focus_task_file, bitable_path, dry_run=args.dry_run) if focus_task_file else {}
    except LarkError as exc:
        if not args.fallback_files:
            raise
        publish_errors.append({"stage": "create_bitable", "error": str(exc)})
        bitable_result = {"ok": False, "error": str(exc), "fallback": "file"}
    doc_url = str(doc_result.get("url") or doc_result.get("remote_url") or "")
    bitable_url = str(bitable_result.get("url") or bitable_result.get("remote_url") or "")
    message_result = {}
    file_message_results: list[dict[str, Any]] = []
    recipient_user_id = args.recipient_user_id
    if args.send:
        if not recipient_user_id:
            recipient_user_id = f"query:{args.recipient}" if args.dry_run else resolve_user_id(args.recipient)
        message_values = {
            "doc_url": doc_url,
            "bitable_url": bitable_url,
            "report_md": str(Path(args.report_md).resolve()),
            "focus_task_file": str(Path(focus_task_file).resolve()) if focus_task_file else "",
            "low_score_csv": str(Path(args.low_score_csv).resolve()) if args.low_score_csv else "",
            "doc_path": doc_path,
            "bitable_path": bitable_path,
        }
        default_message = "\n".join(
            part
            for part in [
                "PaiWork 报告已生成。",
                f"报告：{doc_url}" if doc_url else "",
                f"关注任务表：{bitable_url}" if bitable_url else "",
            ]
            if part
        )
        message = render_message_template(args.message, message_values) if args.message else default_message
        message = append_missing_message_links(message, doc_url=doc_url, bitable_url=bitable_url)
        if publish_errors and args.fallback_files:
            error_summary = "；".join(f"{item['stage']}: {item['error'][:180]}" for item in publish_errors)
            message = "\n".join(
                part
                for part in [
                    message,
                    "",
                    "飞书文档或多维表创建失败，已改为发送本地文件附件。",
                    f"失败原因：{error_summary}" if error_summary else "",
                ]
                if part
            )
        message_result = send_message(recipient_user_id, message, dry_run=args.dry_run, send_as=args.send_as)
        if publish_errors and args.fallback_files:
            file_message_results.append(send_file(recipient_user_id, args.report_md, dry_run=args.dry_run, send_as=args.send_as))
            if focus_task_file:
                file_message_results.append(send_file(recipient_user_id, focus_task_file, dry_run=args.dry_run, send_as=args.send_as))
    payload = {
        "schema_version": "paiobs-lark-publish/v1",
        "report_md": str(Path(args.report_md).resolve()),
        "focus_task_file": str(Path(focus_task_file).resolve()) if focus_task_file else "",
        "low_score_csv": str(Path(args.low_score_csv).resolve()) if args.low_score_csv else "",
        "doc_path": doc_path,
        "bitable_path": bitable_path,
        "doc": doc_result,
        "bitable": bitable_result,
        "send": bool(args.send),
        "recipient": args.recipient,
        "recipient_user_id": recipient_user_id,
        "message": message_result,
        "send_as": args.send_as,
        "fallback_files": bool(args.fallback_files),
        "publish_errors": publish_errors,
        "file_messages": file_message_results,
    }
    paiobs.output_payload(payload, args, default_format=args.format)


def cmd_doctor(args: argparse.Namespace) -> None:
    if args.lark_profile:
        os.environ["PAI_OBS_LARK_PROFILE"] = args.lark_profile
    try:
        prefix, _cwd, _env, mode = lark_cli_invocation()
        lark_cli_error = ""
    except Exception as exc:
        prefix, mode, lark_cli_error = [], "", str(exc)
    checks: dict[str, Any] = {
        "skill_lark_cli": str(SKILL_LARK_CLI),
        "skill_lark_home": str(SKILL_LARK_HOME),
        "lark_cli_prefix": prefix,
        "lark_cli_mode": mode,
        "lark_cli_error": lark_cli_error,
        "lark_profile": args.lark_profile or os.environ.get("PAI_OBS_LARK_PROFILE", ""),
    }
    if os.environ.get("PAI_OBS_FEISHU_SYNC") or os.environ.get("FEISHU_SYNC_BIN") or SKILL_FEISHU_SYNC.is_file() or shutil.which("feishu-sync"):
        try:
            checks["legacy_feishu_sync"] = feishu_sync_bin()
            checks["feishu_sync_doctor"] = feishu_sync_json(["doctor"])
        except Exception as exc:
            checks["legacy_feishu_sync_error"] = str(exc)
    else:
        checks["legacy_feishu_sync"] = "not configured; not required for direct skill-local lark publishing"
    try:
        auth_status = lark_json(["auth", "status", "--verify"])
        scopes = split_scopes(auth_status.get("scope") if isinstance(auth_status, dict) else "")
        checks["lark_auth_status"] = auth_status
        checks["required_send_scopes"] = REQUIRED_SEND_SCOPES
        checks["missing_send_scopes"] = [scope for scope in REQUIRED_SEND_SCOPES if scope not in scopes]
        checks["optional_contact_scopes"] = OPTIONAL_CONTACT_SCOPES
        checks["missing_optional_contact_scopes"] = [scope for scope in OPTIONAL_CONTACT_SCOPES if scope not in scopes]
        checks["send_as_recommendation"] = "bot" if checks["missing_send_scopes"] else "auto"
    except Exception as exc:
        checks["lark_auth_status_error"] = str(exc)
    paiobs.output_payload({"schema_version": "paiobs-lark-doctor/v1", **checks}, args, default_format=args.format)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Publish PaiWork observability reports to Feishu")
    sub = parser.add_subparsers(dest="command", required=True)

    publish = sub.add_parser("publish", help="Create Feishu doc/bitable and optionally DM the recipient")
    publish.add_argument("--report-md", required=True)
    publish.add_argument("--focus-task-file", default="")
    publish.add_argument("--low-score-csv", default="", help="Deprecated compatibility alias for --focus-task-file")
    publish.add_argument("--doc-path", default="")
    publish.add_argument("--bitable-path", default="")
    publish.add_argument("--send", action="store_true")
    publish.add_argument("--recipient", default="fengchao")
    publish.add_argument("--recipient-user-id", default="")
    publish.add_argument("--message", default="")
    publish.add_argument("--lark-profile", default=os.environ.get("PAI_OBS_LARK_PROFILE", DEFAULT_LARK_PROFILE))
    publish.add_argument("--send-as", choices=["auto", "user", "bot"], default=os.environ.get("PAI_OBS_LARK_SEND_AS", "auto"))
    publish.add_argument("--fallback-files", action=argparse.BooleanOptionalAction, default=True)
    publish.add_argument("--dry-run", action="store_true")
    paiobs.add_output_args(publish, default_format="json")
    publish.set_defaults(func=cmd_publish)

    doctor = sub.add_parser("doctor", help="Check local Feishu publishing dependencies")
    doctor.add_argument("--lark-profile", default=os.environ.get("PAI_OBS_LARK_PROFILE", DEFAULT_LARK_PROFILE))
    paiobs.add_output_args(doctor, default_format="json")
    doctor.set_defaults(func=cmd_doctor)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        args.func(args)
        return 0
    except LarkError as exc:
        sys.stderr.write(f"ERROR: {exc}\n")
        return 1
    except KeyboardInterrupt:
        sys.stderr.write("Interrupted\n")
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
